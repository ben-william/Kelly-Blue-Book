from bs4 import BeautifulSoup
import openpyxl
import requests
import time
import pandas as pd

# URL = "https://www.kbb.com/bmw/5-series/2020/m550i-xdrive-sedan-4d/"
# urlList = ['https://www.kbb.com/bmw/5-series/2020/m550i-xdrive-sedan-4d/',
#            'https://www.kbb.com/saab/900/1992/s-convertible-2d/', 'https://www.kbb.com/porsche/911/2021']

headers = {
    "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/97.0.4692.99 Safari/537.36",

    # "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_14_5)AppleWebKit/605.1.15 (KHTML, like Gecko)Version/12.1.1 Safari/605.1.15",
    "Accept-Language": 'en-US',
    "Accept-Encoding": 'gzip',
    "Accept": 'text/html',
    "Referer": 'https://www.google.com',
}

# Import URL List
urlList = []
urlBook = openpyxl.load_workbook(
    '/Users/ben/Desktop/Coding/Python/MMYT Error Rerun.xlsx')
urlSheet = urlBook.active
for x in range(2, urlSheet.max_row):
    urlList.append(urlSheet.cell(row=x, column=1).value)
print(urlList)

# Create Excel

excel = openpyxl.Workbook()
sheet = excel.active
sheet.title = 'KBB MMYT + Images'
sheet.append(['URL', 'Make', 'Model', 'Year', 'Trim',
             'ImageURL'])

checkpoint = 100
complete = 0
badURLs = []
for URL in urlList:
    try:
        result = requests.get(URL, headers=headers)
        doc = BeautifulSoup(result.text, "html.parser")

        # Find Breadcrumb
        breadcrumb = doc.find(
            'div', class_='css-1e2z8je-BreadcrumbContainer edeyoox0')
        # print(breadcrumb)

        # Find MMY links
        MMY_links = breadcrumb.find_all('a')
        # Pull text from all but first link (homepage)
        make = MMY_links[1].text
        model = MMY_links[2].text
        year = MMY_links[3].text
        # print(make, model, year)

        # Find trim text at end of breadcrumb
        trim = breadcrumb.find(
            'div', class_='css-f2dvvi-BreadcrumbCurrentPage edeyoox1').text
        # print(trim)

        # Find image link
        img_div = doc.find('div', id='gallery')
        img_link = img_div.find('a').find('img')['src']
        # print(img_link)

        # Print to sheet
        sheet.append([URL, make, model, year, trim, img_link])

    except:
        print("----- Error ------")
        print(URL)
        badURLs.append(URL)
        print("----- Error ------")

    finally:
        complete += 1
        print(f"Progress: {complete}/{len(urlList)}", "Errors: ",
              len(badURLs))

        # At checkpoint, save and reopen file
        if complete == checkpoint:
            excel.save('MMYT Error Rerun.xlsx')
            excel = openpyxl.load_workbook(filename='MMYT Error Rerun.xlsx')
            sheet = excel.active
            print("Saved & Reopened, Rows: ", sheet.max_row)
            checkpoint += 100
            # Save error progress
            with open('MMYT Errors.csv', mode='w') as file:
                file.write('Bad URLs\n')
                for i in badURLs:
                    file.write(f'{i}\n')

        time.sleep(1.5)


# Save Excel
excel.save('MMYT Error Rerun.xlsx')

# Save Errors
with open('MMYT Errors.csv', mode='w') as file:
    file.write('Bad URLs\n')
    for i in badURLs:
        file.write(f'{i}\n')

#
print('done!')
