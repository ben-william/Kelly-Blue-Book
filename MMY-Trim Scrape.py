from bs4 import BeautifulSoup
import openpyxl
import requests
import time

# Test URL: https://www.kbb.com/audi/a3/2018/

headers = {
    "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/97.0.4692.99 Safari/537.36"
}

# Import URL List
urlList = []
urlBook = openpyxl.load_workbook('KBB_MMY.xlsx', data_only=True)
urlSheet = urlBook.active
for x in range(2, 20000):
    urlList.append(urlSheet.cell(row=x, column=5).value)
print(urlList)
badURLs = []

# Create Excel
excel = openpyxl.Workbook()
sheet = excel.active
sheet.title = 'Trim URLs'
sheet.append(['MMY URL', 'Trim URL'])
complete = 0
checkpoint = 50


for URL in urlList:

    try:
        result = requests.get(URL, headers=headers)
        doc = BeautifulSoup(result.text, "html.parser")

        # find trim cards
        trim_cards = doc.find_all(
            'div', class_='card-heading css-1jfingr erywcq01')

        # find all trim names & links
        for card in trim_cards:
            trim_name = card.find('h3').text
            trim_url = card.parent['href']
            full_trim_url = "https://www.kbb.com" + trim_url
            sheet.append([URL, full_trim_url])

        complete += 1

        # At checkpoint, save and reopen file
        if complete == checkpoint:
            excel.save('KBB_MMY_Trim_URLs.xlsx')
            excel = openpyxl.load_workbook(filename='KBB_MMY_Trim_URLs.xlsx')
            sheet = excel.active
            checkpoint += 50

        time.sleep(1.5)
    except:
        print("----- Error ------")
        badURLs.append(URL)
        print("----- Error ------")

# Save Excel
excel.save('KBB_MMY_Trim_URLs.xlsx')
